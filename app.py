import streamlit as st
import pandas as pd
import openpyxl
import io
import json
from datetime import datetime, date
import re

# ───────────────────────────────────────────
# 페이지 설정
# ───────────────────────────────────────────
st.set_page_config(
    page_title="CTR 선행품질 대시보드",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ───────────────────────────────────────────
# 고급 CSS 스타일
# ───────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700&family=DM+Sans:wght@300;400;500;600&display=swap');

* { box-sizing: border-box; }

.stApp {
    background: #0a0e1a;
    color: #e2e8f0;
    font-family: 'DM Sans', 'Noto Sans KR', sans-serif;
}

.main .block-container {
    padding: 1.5rem 2rem 3rem;
    max-width: 1400px;
}

/* 헤더 */
.dash-header {
    background: linear-gradient(135deg, #0f172a 0%, #1e293b 50%, #0f172a 100%);
    border: 1px solid #334155;
    border-radius: 16px;
    padding: 28px 36px;
    margin-bottom: 24px;
    position: relative;
    overflow: hidden;
}
.dash-header::before {
    content: '';
    position: absolute;
    top: -60px; right: -60px;
    width: 200px; height: 200px;
    background: radial-gradient(circle, rgba(59,130,246,0.15) 0%, transparent 70%);
    border-radius: 50%;
}
.dash-header::after {
    content: '';
    position: absolute;
    bottom: -40px; left: 40%;
    width: 300px; height: 150px;
    background: radial-gradient(ellipse, rgba(99,102,241,0.08) 0%, transparent 70%);
}
.header-title {
    font-size: 1.8rem;
    font-weight: 600;
    color: #f1f5f9;
    margin: 0 0 6px;
    letter-spacing: -0.5px;
}
.header-sub {
    font-size: 0.9rem;
    color: #64748b;
    margin: 0;
}
.header-badge {
    display: inline-block;
    background: rgba(59,130,246,0.15);
    color: #60a5fa;
    border: 1px solid rgba(59,130,246,0.3);
    border-radius: 20px;
    padding: 3px 12px;
    font-size: 0.75rem;
    font-weight: 500;
    margin-bottom: 10px;
}

/* KPI 카드 */
.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 14px;
    margin-bottom: 20px;
}
.kpi-card {
    background: #111827;
    border: 1px solid #1f2937;
    border-radius: 14px;
    padding: 20px 22px;
    position: relative;
    overflow: hidden;
    transition: border-color 0.2s;
}
.kpi-card:hover { border-color: #374151; }
.kpi-label {
    font-size: 0.72rem;
    font-weight: 500;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    margin-bottom: 10px;
}
.kpi-value {
    font-size: 2.4rem;
    font-weight: 600;
    line-height: 1;
    margin-bottom: 6px;
    font-family: 'DM Sans', sans-serif;
}
.kpi-desc {
    font-size: 0.75rem;
    color: #4b5563;
}
.kpi-green  { border-left: 3px solid #22c55e; }
.kpi-yellow { border-left: 3px solid #f59e0b; }
.kpi-red    { border-left: 3px solid #ef4444; }
.kpi-blue   { border-left: 3px solid #3b82f6; }
.kpi-green  .kpi-label { color: #4ade80; }
.kpi-yellow .kpi-label { color: #fbbf24; }
.kpi-red    .kpi-label { color: #f87171; }
.kpi-blue   .kpi-label { color: #60a5fa; }
.kpi-green  .kpi-value { color: #4ade80; }
.kpi-yellow .kpi-value { color: #fbbf24; }
.kpi-red    .kpi-value { color: #f87171; }
.kpi-blue   .kpi-value { color: #93c5fd; }

/* 섹션 카드 */
.section-card {
    background: #111827;
    border: 1px solid #1f2937;
    border-radius: 14px;
    padding: 20px 22px;
    margin-bottom: 16px;
}
.section-title {
    font-size: 0.85rem;
    font-weight: 600;
    color: #94a3b8;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    margin-bottom: 16px;
    padding-bottom: 10px;
    border-bottom: 1px solid #1f2937;
}

/* 진행률 바 */
.progress-row {
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 12px;
}
.progress-label {
    font-size: 0.8rem;
    color: #94a3b8;
    min-width: 140px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.progress-track {
    flex: 1;
    height: 6px;
    background: #1f2937;
    border-radius: 3px;
    overflow: hidden;
}
.progress-fill-g { background: #22c55e; height: 6px; border-radius: 3px; }
.progress-fill-y { background: #f59e0b; height: 6px; border-radius: 3px; }
.progress-fill-r { background: #ef4444; height: 6px; border-radius: 3px; }
.progress-pct {
    font-size: 0.78rem;
    font-weight: 600;
    min-width: 38px;
    text-align: right;
}

/* 이슈 테이블 */
.issue-table {
    width: 100%;
    border-collapse: collapse;
    font-size: 0.82rem;
}
.issue-table th {
    background: #0f172a;
    color: #64748b;
    font-weight: 500;
    font-size: 0.72rem;
    letter-spacing: 0.06em;
    text-transform: uppercase;
    padding: 10px 12px;
    text-align: left;
    border-bottom: 1px solid #1f2937;
}
.issue-table td {
    padding: 11px 12px;
    border-bottom: 1px solid #1a2233;
    color: #cbd5e1;
    vertical-align: middle;
}
.issue-table tr:hover td { background: #161f2e; }
.badge-g { background: rgba(34,197,94,0.15); color: #4ade80; border: 1px solid rgba(34,197,94,0.3); border-radius: 20px; padding: 2px 10px; font-size: 0.72rem; font-weight: 600; }
.badge-y { background: rgba(245,158,11,0.15); color: #fbbf24; border: 1px solid rgba(245,158,11,0.3); border-radius: 20px; padding: 2px 10px; font-size: 0.72rem; font-weight: 600; }
.badge-r { background: rgba(239,68,68,0.15); color: #f87171; border: 1px solid rgba(239,68,68,0.3); border-radius: 20px; padding: 2px 10px; font-size: 0.72rem; font-weight: 600; }
.badge-none { background: rgba(100,116,139,0.15); color: #64748b; border: 1px solid rgba(100,116,139,0.3); border-radius: 20px; padding: 2px 10px; font-size: 0.72rem; }

/* 고객사 태그 */
.tag-tesla  { background: rgba(239,68,68,0.15); color: #f87171; border-radius: 6px; padding: 2px 8px; font-size: 0.75rem; font-weight: 600; }
.tag-renault{ background: rgba(99,102,241,0.15); color: #a5b4fc; border-radius: 6px; padding: 2px 8px; font-size: 0.75rem; font-weight: 600; }
.tag-nexteer{ background: rgba(20,184,166,0.15); color: #2dd4bf; border-radius: 6px; padding: 2px 8px; font-size: 0.75rem; font-weight: 600; }
.tag-other  { background: rgba(100,116,139,0.15); color: #94a3b8; border-radius: 6px; padding: 2px 8px; font-size: 0.75rem; font-weight: 600; }

/* 업로드 영역 */
.upload-zone {
    background: #0f172a;
    border: 2px dashed #334155;
    border-radius: 16px;
    padding: 48px;
    text-align: center;
    margin: 40px 0;
}
.upload-icon { font-size: 3rem; margin-bottom: 16px; }
.upload-title { font-size: 1.2rem; font-weight: 600; color: #94a3b8; margin-bottom: 8px; }
.upload-desc { font-size: 0.85rem; color: #475569; }

/* 필터 바 */
.filter-bar {
    display: flex;
    gap: 8px;
    margin-bottom: 16px;
    flex-wrap: wrap;
}

/* Streamlit 컴포넌트 덮어쓰기 */
.stSelectbox > div > div { background: #111827 !important; border-color: #1f2937 !important; color: #e2e8f0 !important; }
.stFileUploader { background: transparent !important; }
div[data-testid="stFileUploader"] { background: #0f172a; border: 2px dashed #334155; border-radius: 16px; padding: 20px; }
.stButton > button {
    background: #1e3a5f !important;
    color: #60a5fa !important;
    border: 1px solid #2563eb !important;
    border-radius: 8px !important;
    font-weight: 500 !important;
}
.stButton > button:hover { background: #2563eb !important; color: white !important; }
.stTabs [data-baseweb="tab"] { background: transparent !important; color: #64748b !important; }
.stTabs [aria-selected="true"] { color: #60a5fa !important; border-bottom: 2px solid #3b82f6 !important; }
h1,h2,h3 { color: #f1f5f9 !important; }

/* 새 이슈 입력 폼 */
.form-card {
    background: #0f172a;
    border: 1px solid #334155;
    border-radius: 12px;
    padding: 20px;
    margin-top: 12px;
}
</style>
""", unsafe_allow_html=True)


# ───────────────────────────────────────────
# 헬퍼 함수
# ───────────────────────────────────────────
def customer_tag(name):
    if name is None: return '<span class="tag-other">-</span>'
    n = str(name).upper()
    if 'TESLA' in n: return f'<span class="tag-tesla">{name}</span>'
    if 'RENAULT' in n or 'RG' in n: return f'<span class="tag-renault">{name}</span>'
    if 'NEXTEER' in n: return f'<span class="tag-nexteer">{name}</span>'
    return f'<span class="tag-other">{name}</span>'

def status_badge(v):
    if v is None or str(v).strip() == '': return '<span class="badge-none">미입력</span>'
    v = str(v).strip().upper()
    if 'G' in v: return '<span class="badge-g">● 완료</span>'
    if 'Y' in v: return '<span class="badge-y">● 진행중</span>'
    if 'R' in v: return '<span class="badge-r">● 미완료</span>'
    return f'<span class="badge-none">{v}</span>'

def safe_str(v):
    if v is None: return ''
    return str(v).strip()

def parse_excel(file_bytes):
    """엑셀 파일 파싱"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), keep_vba=True, data_only=True)
    data = {}

    # OPEN ISSUE LIST
    if '5.OPEN ISSUE LIST' in wb.sheetnames:
        ws = wb['5.OPEN ISSUE LIST']
        issues = []
        for row in ws.iter_rows(min_row=9, values_only=True):
            vals = list(row)
            # No, 개발단계, 프로그램(고객사), 품명, 품목, 품번, 이슈내용, ...완료여부
            no_val = vals[1] if len(vals) > 1 else None
            if no_val is None: continue
            # HYPERLINK 수식에서 숫자 추출
            no_str = str(no_val) if no_val else ''
            if 'HYPERLINK' in no_str:
                m = re.search(r'"(\d+)"', no_str)
                no_str = m.group(1) if m else no_str
            stage    = vals[2] if len(vals) > 2 else None
            customer = vals[3] if len(vals) > 3 else None
            part_name= vals[4] if len(vals) > 4 else None
            part_type= vals[5] if len(vals) > 5 else None
            part_no  = vals[6] if len(vals) > 6 else None
            issue    = vals[7] if len(vals) > 7 else None
            action   = vals[11] if len(vals) > 11 else None
            start_dt = vals[13] if len(vals) > 13 else None
            target_dt= vals[14] if len(vals) > 14 else None
            done_dt  = vals[15] if len(vals) > 15 else None
            status   = vals[15] if len(vals) > 15 else None  # 완료여부

            if issue or action:
                issues.append({
                    'no': no_str,
                    'stage': safe_str(stage),
                    'customer': safe_str(customer),
                    'part_name': safe_str(part_name),
                    'part_no': safe_str(part_no),
                    'issue': safe_str(issue),
                    'action': safe_str(action),
                    'target_dt': safe_str(target_dt),
                    'status': safe_str(status),
                })
        data['issues'] = issues

    # 고객 요구사항
    if '5.고객 요구사항 및대응' in wb.sheetnames:
        ws = wb['5.고객 요구사항 및대응']
        reqs = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = list(row)
            no = vals[1] if len(vals) > 1 else None
            if no is None: continue
            reqs.append({
                'no': safe_str(no),
                'date': safe_str(vals[2] if len(vals)>2 else None),
                'stage': safe_str(vals[3] if len(vals)>3 else None),
                'customer': safe_str(vals[4] if len(vals)>4 else None),
                'issue': safe_str(vals[5] if len(vals)>5 else None),
                'status': safe_str(vals[6] if len(vals)>6 else None),
                'action': safe_str(vals[7] if len(vals)>7 else None),
                'target_dt': safe_str(vals[8] if len(vals)>8 else None),
                'assignee': safe_str(vals[10] if len(vals)>10 else None),
                'done': safe_str(vals[12] if len(vals)>12 else None),
            })
        data['requirements'] = reqs

    # 고객 감사
    if '6.고객방문&감사 대응' in wb.sheetnames:
        ws = wb['6.고객방문&감사 대응']
        audits = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = list(row)
            no = vals[1] if len(vals) > 1 else None
            if no is None: continue
            audits.append({
                'no': safe_str(no),
                'date': safe_str(vals[2] if len(vals)>2 else None),
                'stage': safe_str(vals[3] if len(vals)>3 else None),
                'customer': safe_str(vals[4] if len(vals)>4 else None),
                'issue': safe_str(vals[5] if len(vals)>5 else None),
                'status': safe_str(vals[6] if len(vals)>6 else None),
                'action': safe_str(vals[7] if len(vals)>7 else None),
                'target_dt': safe_str(vals[8] if len(vals)>8 else None),
                'assignee': safe_str(vals[10] if len(vals)>10 else None),
                'done': safe_str(vals[12] if len(vals)>12 else None),
            })
        data['audits'] = audits

    return data


def count_status(items, key='status'):
    g = sum(1 for i in items if 'G' in str(i.get(key,'')).upper())
    y = sum(1 for i in items if 'Y' in str(i.get(key,'')).upper() and 'G' not in str(i.get(key,'')).upper())
    r = sum(1 for i in items if 'R' in str(i.get(key,'')).upper())
    return g, y, r


# ───────────────────────────────────────────
# 헤더
# ───────────────────────────────────────────
st.markdown("""
<div class="dash-header">
  <div class="header-badge">CTR 선행개발품질팀</div>
  <div class="header-title">OPEN ISSUE 관리 대시보드</div>
  <div class="header-sub">엑셀 파일을 업로드하면 실시간으로 현황을 분석합니다</div>
</div>
""", unsafe_allow_html=True)


# ───────────────────────────────────────────
# 세션 상태
# ───────────────────────────────────────────
if 'data' not in st.session_state:
    st.session_state.data = None
if 'manual_issues' not in st.session_state:
    st.session_state.manual_issues = []


# ───────────────────────────────────────────
# 파일 업로드
# ───────────────────────────────────────────
col_up, col_btn = st.columns([4, 1])
with col_up:
    uploaded = st.file_uploader(
        "OPEN_ISSUE 엑셀 파일 업로드 (.xlsm, .xlsx)",
        type=["xlsm", "xlsx"],
        label_visibility="collapsed"
    )
with col_btn:
    if uploaded and st.button("📊 분석 시작", use_container_width=True):
        with st.spinner("파일 분석 중..."):
            st.session_state.data = parse_excel(uploaded.read())
        st.rerun()

# 자동 분석
if uploaded and st.session_state.data is None:
    with st.spinner("파일 분석 중..."):
        st.session_state.data = parse_excel(uploaded.read())
    st.rerun()


# ───────────────────────────────────────────
# 데이터 없을 때
# ───────────────────────────────────────────
if st.session_state.data is None:
    st.markdown("""
    <div class="upload-zone">
      <div class="upload-icon">📂</div>
      <div class="upload-title">OPEN_ISSUE 엑셀 파일을 업로드하세요</div>
      <div class="upload-desc">xlsm / xlsx 파일을 드래그하거나 위 버튼을 클릭하세요<br>파일을 올리면 자동으로 대시보드가 생성됩니다</div>
    </div>
    """, unsafe_allow_html=True)

    # 수동 입력 섹션
    st.markdown('<div class="section-card"><div class="section-title">✏️ 직접 이슈 입력</div>', unsafe_allow_html=True)
    with st.form("manual_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            m_customer = st.selectbox("고객사", ["TESLA PM3", "RENAULT", "NEXTEER", "기타"])
            m_stage = st.selectbox("개발단계", ["1.제품기획", "2.제품설계", "3.공정설계", "4.공정유효성평가", "5.고객대응", "6.고객감사"])
        with c2:
            m_part = st.text_input("품명")
            m_issue = st.text_input("이슈내용")
        with c3:
            m_action = st.text_input("개선대책")
            m_status = st.selectbox("상태", ["R (미완료)", "Y (진행중)", "G (완료)"])

        if st.form_submit_button("이슈 추가"):
            st.session_state.manual_issues.append({
                'no': str(len(st.session_state.manual_issues)+1),
                'stage': m_stage,
                'customer': m_customer,
                'part_name': m_part,
                'part_no': '-',
                'issue': m_issue,
                'action': m_action,
                'target_dt': '',
                'status': m_status[0],
            })
            st.success("이슈가 추가됐어요!")
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.manual_issues:
        st.session_state.data = {'issues': st.session_state.manual_issues, 'requirements': [], 'audits': []}
        st.rerun()

    st.stop()


# ───────────────────────────────────────────
# 데이터 있을 때 — 대시보드
# ───────────────────────────────────────────
data = st.session_state.data
issues = data.get('issues', []) + st.session_state.manual_issues
reqs   = data.get('requirements', [])
audits = data.get('audits', [])

# 통계
g_i, y_i, r_i = count_status(issues)
g_r, y_r, r_r = count_status(reqs, 'done')
g_a, y_a, r_a = count_status(audits, 'done')

total_issues = len([i for i in issues if i.get('issue')])
total_reqs   = len([i for i in reqs if i.get('issue')])
total_audits = len([i for i in audits if i.get('issue')])
total_all = total_issues + total_reqs + total_audits
g_all = g_i + g_r + g_a
y_all = y_i + y_r + y_a
r_all = r_i + r_r + r_a

# ── KPI 카드 ──────────────────────────────
st.markdown(f"""
<div class="kpi-grid">
  <div class="kpi-card kpi-blue">
    <div class="kpi-label">전체 이슈</div>
    <div class="kpi-value">{total_all}</div>
    <div class="kpi-desc">OPEN ISSUE + 고객요구 + 감사</div>
  </div>
  <div class="kpi-card kpi-green">
    <div class="kpi-label">완료 🟢</div>
    <div class="kpi-value">{g_all}</div>
    <div class="kpi-desc">G 완료 건수</div>
  </div>
  <div class="kpi-card kpi-yellow">
    <div class="kpi-label">진행중 🟡</div>
    <div class="kpi-value">{y_all}</div>
    <div class="kpi-desc">Y 진행중 건수</div>
  </div>
  <div class="kpi-card kpi-red">
    <div class="kpi-label">미완료 🔴</div>
    <div class="kpi-value">{r_all}</div>
    <div class="kpi-desc">R 미완료 건수</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── 탭 ────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs(["📋 OPEN ISSUE LIST", "📞 고객 요구사항", "🔍 고객 감사", "➕ 이슈 직접 입력"])

# ── 탭1: OPEN ISSUE LIST ──────────────────
with tab1:
    col_a, col_b = st.columns([2, 1])

    with col_a:
        # 단계별 진행률
        stage_map = {}
        for iss in issues:
            s = iss.get('stage','기타')
            if not s: s = '기타'
            if s not in stage_map:
                stage_map[s] = {'g':0,'y':0,'r':0,'n':0}
            st_val = str(iss.get('status','')).upper()
            if 'G' in st_val: stage_map[s]['g'] += 1
            elif 'Y' in st_val: stage_map[s]['y'] += 1
            elif 'R' in st_val: stage_map[s]['r'] += 1
            else: stage_map[s]['n'] += 1

        bars_html = '<div class="section-card"><div class="section-title">APQP 단계별 이슈 현황</div>'
        for stage, cnt in stage_map.items():
            total_s = cnt['g'] + cnt['y'] + cnt['r'] + cnt['n']
            if total_s == 0: continue
            pct = int(cnt['g'] / total_s * 100)
            color = 'g' if pct >= 70 else ('y' if pct >= 30 else 'r')
            txt_color = '#4ade80' if pct >= 70 else ('#fbbf24' if pct >= 30 else '#f87171')
            label = stage.replace('4.공정유효성평가&양산이관','4.공정유효성').replace('5.고객 요구사항 및대응','5.고객대응')
            bars_html += f"""
            <div class="progress-row">
              <span class="progress-label" title="{stage}">{label[:18]}</span>
              <div class="progress-track"><div class="progress-fill-{color}" style="width:{pct}%"></div></div>
              <span class="progress-pct" style="color:{txt_color}">{pct}%</span>
            </div>"""
        bars_html += '</div>'
        st.markdown(bars_html, unsafe_allow_html=True)

    with col_b:
        # 고객사별 이슈
        cust_map = {}
        for iss in issues:
            c = iss.get('customer','기타') or '기타'
            cust_map[c] = cust_map.get(c, 0) + 1

        cust_html = '<div class="section-card"><div class="section-title">고객사별 이슈</div>'
        max_cnt = max(cust_map.values()) if cust_map else 1
        colors = {'TESLA PM3':'#ef4444','RENAULT':'#818cf8','NEXTEER':'#2dd4bf'}
        for cust, cnt in sorted(cust_map.items(), key=lambda x:-x[1]):
            pct = int(cnt / max_cnt * 100)
            color = colors.get(cust, '#64748b')
            cust_html += f"""
            <div style="margin-bottom:12px">
              <div style="display:flex;justify-content:space-between;font-size:0.8rem;margin-bottom:4px">
                <span style="color:#94a3b8">{cust}</span>
                <span style="color:{color};font-weight:600">{cnt}건</span>
              </div>
              <div style="background:#1f2937;border-radius:3px;height:6px">
                <div style="background:{color};width:{pct}%;height:6px;border-radius:3px"></div>
              </div>
            </div>"""
        cust_html += '</div>'
        st.markdown(cust_html, unsafe_allow_html=True)

    # 이슈 테이블
    filter_col1, filter_col2 = st.columns(2)
    with filter_col1:
        status_filter = st.selectbox("상태 필터", ["전체","🔴 미완료(R)","🟡 진행중(Y)","🟢 완료(G)"], key="sf1")
    with filter_col2:
        search = st.text_input("이슈 검색", placeholder="키워드 입력...", key="sr1")

    filtered = issues
    if status_filter == "🔴 미완료(R)":
        filtered = [i for i in issues if 'R' in str(i.get('status','')).upper()]
    elif status_filter == "🟡 진행중(Y)":
        filtered = [i for i in issues if 'Y' in str(i.get('status','')).upper()]
    elif status_filter == "🟢 완료(G)":
        filtered = [i for i in issues if 'G' in str(i.get('status','')).upper()]
    if search:
        filtered = [i for i in filtered if search.lower() in str(i.get('issue','')).lower() or search.lower() in str(i.get('customer','')).lower()]

    rows_html = ""
    for iss in filtered[:50]:
        rows_html += f"""<tr>
          <td style="color:#64748b;font-size:0.75rem">{iss.get('no','')}</td>
          <td>{customer_tag(iss.get('customer'))}</td>
          <td style="max-width:100px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;color:#94a3b8;font-size:0.75rem" title="{iss.get('stage','')}">{iss.get('stage','')[:20]}</td>
          <td style="color:#e2e8f0;font-weight:500">{iss.get('issue','')[:40]}</td>
          <td style="color:#94a3b8">{iss.get('action','')[:30]}</td>
          <td style="color:#64748b;font-size:0.75rem">{iss.get('target_dt','')}</td>
          <td>{status_badge(iss.get('status',''))}</td>
        </tr>"""

    table_html = f"""
    <div class="section-card">
      <div class="section-title">OPEN ISSUE LIST ({len(filtered)}건)</div>
      <div style="overflow-x:auto">
      <table class="issue-table">
        <thead><tr>
          <th width="40">NO</th><th width="110">고객사</th><th width="130">개발단계</th>
          <th>이슈내용</th><th>개선대책</th><th width="100">목표일자</th><th width="90">상태</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
      </div>
    </div>"""
    st.markdown(table_html, unsafe_allow_html=True)

# ── 탭2: 고객 요구사항 ────────────────────
with tab2:
    if not reqs or not any(i.get('issue') for i in reqs):
        st.markdown('<div class="section-card"><div style="text-align:center;color:#475569;padding:40px">데이터가 없습니다. 엑셀에 고객 요구사항 내용을 입력해주세요.</div></div>', unsafe_allow_html=True)
    else:
        rows_html = ""
        for r in reqs:
            if not r.get('issue'): continue
            rows_html += f"""<tr>
              <td style="color:#64748b">{r.get('no','')}</td>
              <td style="color:#94a3b8;font-size:0.75rem">{r.get('date','')}</td>
              <td>{customer_tag(r.get('customer'))}</td>
              <td style="color:#94a3b8;font-size:0.75rem">{r.get('stage','')}</td>
              <td style="color:#e2e8f0">{r.get('issue','')[:40]}</td>
              <td style="color:#94a3b8">{r.get('action','')[:30]}</td>
              <td style="color:#64748b;font-size:0.75rem">{r.get('target_dt','')}</td>
              <td style="color:#94a3b8">{r.get('assignee','')}</td>
              <td>{status_badge(r.get('done',''))}</td>
            </tr>"""
        st.markdown(f"""
        <div class="section-card">
          <div class="section-title">고객 요구사항 대응 LIST</div>
          <div style="overflow-x:auto">
          <table class="issue-table">
            <thead><tr>
              <th>NO</th><th>요청일자</th><th>고객사</th><th>단계</th>
              <th>요청내용</th><th>개선대책</th><th>목표일자</th><th>담당자</th><th>종결</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
          </div>
        </div>""", unsafe_allow_html=True)

# ── 탭3: 고객 감사 ────────────────────────
with tab3:
    if not audits or not any(i.get('issue') for i in audits):
        st.markdown('<div class="section-card"><div style="text-align:center;color:#475569;padding:40px">데이터가 없습니다. 엑셀에 고객 감사 내용을 입력해주세요.</div></div>', unsafe_allow_html=True)
    else:
        rows_html = ""
        for a in audits:
            if not a.get('issue'): continue
            rows_html += f"""<tr>
              <td style="color:#64748b">{a.get('no','')}</td>
              <td style="color:#94a3b8;font-size:0.75rem">{a.get('date','')}</td>
              <td>{customer_tag(a.get('customer'))}</td>
              <td style="color:#94a3b8;font-size:0.75rem">{a.get('stage','')}</td>
              <td style="color:#e2e8f0">{a.get('issue','')[:40]}</td>
              <td style="color:#94a3b8">{a.get('action','')[:30]}</td>
              <td style="color:#64748b;font-size:0.75rem">{a.get('target_dt','')}</td>
              <td style="color:#94a3b8">{a.get('assignee','')}</td>
              <td>{status_badge(a.get('done',''))}</td>
            </tr>"""
        st.markdown(f"""
        <div class="section-card">
          <div class="section-title">고객 감사 OPEN ISSUE LIST</div>
          <div style="overflow-x:auto">
          <table class="issue-table">
            <thead><tr>
              <th>NO</th><th>방문일자</th><th>고객사</th><th>단계</th>
              <th>요청내용</th><th>개선대책</th><th>목표일자</th><th>담당자</th><th>종결</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
          </table>
          </div>
        </div>""", unsafe_allow_html=True)

# ── 탭4: 직접 입력 ────────────────────────
with tab4:
    st.markdown('<div class="section-card"><div class="section-title">➕ 새 이슈 직접 입력</div>', unsafe_allow_html=True)
    with st.form("add_issue_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            new_customer = st.selectbox("고객사", ["TESLA PM3","RENAULT","NEXTEER","기타"])
            new_stage = st.selectbox("개발단계", ["1.제품기획","2.제품설계","3.공정설계","4.공정유효성평가","5.고객대응","6.고객감사"])
        with c2:
            new_part = st.text_input("품명")
            new_issue = st.text_area("이슈내용", height=80)
        with c3:
            new_action = st.text_area("개선대책", height=80)
            new_status = st.selectbox("현재 상태", ["R (미완료)","Y (진행중)","G (완료)"])

        submitted = st.form_submit_button("✅ 이슈 추가", use_container_width=True)
        if submitted:
            new_item = {
                'no': str(len(st.session_state.manual_issues)+1),
                'stage': new_stage,
                'customer': new_customer,
                'part_name': new_part,
                'part_no': '-',
                'issue': new_issue,
                'action': new_action,
                'target_dt': '',
                'status': new_status[0],
            }
            st.session_state.manual_issues.append(new_item)
            if st.session_state.data is not None:
                st.session_state.data['issues'].append(new_item)
            st.success(f"이슈가 추가됐어요! ({new_customer} - {new_issue[:20]}...)")
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

    if st.session_state.manual_issues:
        st.markdown(f'<div class="section-card"><div class="section-title">수동 입력 이슈 ({len(st.session_state.manual_issues)}건)</div>', unsafe_allow_html=True)
        for i, m in enumerate(st.session_state.manual_issues):
            st.markdown(f"""<div style="display:flex;align-items:center;gap:12px;padding:8px 0;border-bottom:1px solid #1f2937">
              <span style="color:#475569;font-size:0.75rem">{m['no']}</span>
              {customer_tag(m['customer'])}
              <span style="color:#e2e8f0;flex:1">{m['issue']}</span>
              {status_badge(m['status'])}
            </div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

        if st.button("🗑️ 수동 입력 이슈 전체 삭제"):
            st.session_state.manual_issues = []
            st.rerun()

# ── 마지막 업데이트 ──────────────────────
st.markdown(f"""
<div style="text-align:center;color:#1f2937;font-size:0.75rem;margin-top:24px;padding-top:16px;border-top:1px solid #1f2937">
  CTR 선행개발품질팀 · OPEN ISSUE 대시보드 · 마지막 갱신: {datetime.now().strftime('%Y-%m-%d %H:%M')}
</div>
""", unsafe_allow_html=True)
